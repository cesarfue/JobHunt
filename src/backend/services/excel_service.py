import openpyxl
from config import Config
from openpyxl.styles import Font


def add_to_excel(company, platform, job_title, url, date):
    try:
        wb = openpyxl.load_workbook(Config.EXCEL_FILE)
        ws = wb["suivi"]

        data = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(row):  # Skip empty rows
                data.append(list(row))

        new_row_data = ["A faire", "CDI", company, platform, job_title, url, date, ""]
        data.append(new_row_data)

        data_sorted = sorted(data, key=lambda x: x[0] or "")

        ws.delete_rows(2, ws.max_row - 1)

        for row_data in data_sorted:
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.font = Font(name="Roboto", size=11)

        for cell in ws[1]:
            cell.font = Font(name="Roboto", size=11, bold=True)

        wb.save(Config.EXCEL_FILE)
        wb.close()

    except Exception as e:
        print(f"Error adding to Excel: {e}")
        raise
